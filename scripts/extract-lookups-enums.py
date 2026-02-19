"""
Extract all enum values from Ocean ACDC Schema Lookups tab and generate
per-table markdown context files in the mapping-skills enums directory.

Follows the INVESTOR-ENUMS.md format. After running, re-run:
  npx tsx scripts/import-servicemac-domain.ts

Usage: python3 scripts/extract-lookups-enums.py
"""

import os
from collections import defaultdict
from datetime import date
from openpyxl import load_workbook

XLSX_PATH = "/Users/grantlee/Dev/mapping-skills/servicemac-m1/Ocean ACDC Schema (go_acdc).xlsx"
ENUMS_DIR = "/Users/grantlee/Dev/mapping-skills/.claude/skills/servicemac-domain/enums"
TODAY = date.today().isoformat()

# Fields with more than this many values get summarized instead of full listing
FULL_LISTING_THRESHOLD = 100


def slugify(name: str) -> str:
    """Convert table name to filename slug: LoanInfo -> LOANINFO"""
    return name.upper().replace(" ", "-").replace(",", "")


def table_label(name: str) -> str:
    """Convert table name to readable label."""
    # Handle multi-table fields like "StopsFlagsAndIndicators, Arm"
    return name.replace(",", " /")


def extract_lookups() -> dict:
    """Read Lookups tab columns A-D, group by ExtractType (table)."""
    wb = load_workbook(XLSX_PATH, read_only=True)
    ws = wb["Lookups"]

    # {table_name: {field_name: [(value, definition), ...]}}
    tables: dict[str, dict[str, list[tuple[str, str | None]]]] = defaultdict(
        lambda: defaultdict(list)
    )

    for row in ws.iter_rows(min_row=2, max_col=4, values_only=True):
        field_name, extract_type, field_value, field_def = row
        if not field_name or field_value is None:
            continue

        fn = str(field_name).strip()
        et = str(extract_type).strip() if extract_type else "Unknown"
        fv = str(field_value).strip()
        fd = str(field_def).strip() if field_def else None

        # Normalize multi-table extract types to primary table
        primary_table = et.split(",")[0].strip()
        tables[primary_table][fn].append((fv, fd))

    wb.close()
    return dict(tables)


def generate_markdown(table_name: str, fields: dict) -> str:
    """Generate markdown for one table's enum values."""
    lines = [
        f"# {table_label(table_name)} Table Enum Values",
        "",
        "**AUTHORITATIVE SOURCE** - Extracted from Ocean ACDC Schema Lookups tab",
        "",
        "```",
        "servicemac-m1/Ocean ACDC Schema (go_acdc).xlsx → Lookups tab",
        "```",
        "",
        f"**Extracted:** {TODAY}",
        "",
        "---",
        "",
    ]

    # Sort fields alphabetically
    for field_name in sorted(fields.keys()):
        values = fields[field_name]
        lines.append(f"## {field_name}")
        lines.append("")
        lines.append(f"**{len(values)} values**")
        lines.append("")

        if len(values) > FULL_LISTING_THRESHOLD:
            # Summarize large fields
            lines.append(
                f"> Too many values to list inline ({len(values)}). "
                f"Query the Lookups tab filtering FieldName = `{field_name}`."
            )
            lines.append("")
            # Show first 10 as sample
            lines.append("**Sample values:**")
            lines.append("")
            lines.append("| Code | Definition |")
            lines.append("|------|------------|")
            for val, defn in values[:10]:
                lines.append(f"| `{val}` | {defn or '-'} |")
            lines.append(f"| ... | *({len(values) - 10} more)* |")
        else:
            # Full listing
            lines.append("| Code | Definition |")
            lines.append("|------|------------|")
            for val, defn in values:
                lines.append(f"| `{val}` | {defn or '-'} |")

        lines.append("")
        lines.append("---")
        lines.append("")

    return "\n".join(lines)


def main():
    os.makedirs(ENUMS_DIR, exist_ok=True)

    tables = extract_lookups()

    total_fields = 0
    total_values = 0
    files_written = 0

    for table_name in sorted(tables.keys()):
        fields = tables[table_name]
        field_count = len(fields)
        value_count = sum(len(v) for v in fields.values())
        total_fields += field_count
        total_values += value_count

        filename = f"{slugify(table_name)}-ENUMS.md"
        filepath = os.path.join(ENUMS_DIR, filename)

        md = generate_markdown(table_name, fields)

        with open(filepath, "w") as f:
            f.write(md)

        files_written += 1
        print(f"  {filename}: {field_count} fields, {value_count} values")

    print(f"\nWrote {files_written} enum files to {ENUMS_DIR}")
    print(f"Total: {total_fields} fields, {total_values} values")
    print(f"\nNext: re-run  npx tsx scripts/import-servicemac-domain.ts")


if __name__ == "__main__":
    main()
